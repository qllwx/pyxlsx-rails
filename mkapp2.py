# -*- coding: utf-8 -*-
from __future__ import unicode_literals
import openpyxl,os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.comments import Comment
import sys,types
from pypinyin import pinyin, lazy_pinyin, Style
reload(sys)
sys.setdefaultencoding('utf-8')
fw=open("makeapp.bash","w")
fy=open("en.yml","w")
fy.write("en:\n")
fw_create=open("Create_data.rb","w")
f_index=open("index.html.erb","w")
dirs=os.listdir('.')
f_index.write("<h1>")
views=[]
dic_type={'long':    'integer',
          'str':     'text',
          'float':   'numeric',
          'unicode': 'string',
          'NoneType':'string'
         }
for f in dirs:
   if f.split(".")[-1]=="xlsx":
       appname=f.split(".")[0]
       f_index.write(appname+"!</h1>\n")
       app_py=lazy_pinyin(appname)
       py=""
       for p in app_py:
           py=py + p
       print appname, py
       fy.write("\n  "+py+": '"+appname+"'")
       fw.write("test -d "+py +"  && rm -fr "+py+"\n")

       fw.write("rails new "+py+"  -f \n")
       fw.write("cd "+py+"\n")
       fw.write("rails generate controller Home index\n")
       app_py=py
       wb = load_workbook(filename = f)
       sheetlist=[]
       for sheet in wb:
           sheetlist.append(sheet.title)
       for sheet in wb:
           sheetname=sheet.title
           sheet_py=lazy_pinyin(sheetname,errors='replace')
           py=""
           for p in sheet_py:
               py=py + p
           print " ",sheetname, py
           sheetname_py=py
           views.append(py+"s")
           fy.write("\n  "+py+": '"+sheetname+"'")
           fw.write("rails generate scaffold "+py)
           f_index.write("<%= link_to '")
           f_index.write(sheetname)
           f_index.write("', controller: '")
           f_index.write(py)
           f_index.write("s' %>\n")
           ws=wb[sheetname]
           row_range=ws[2:2]
           ts=[]
           for cells in row_range:
               s=str(type(cells.value))
               t_s=s[7:len(s)-2]
               ts.append(dic_type.get(t_s,"string"))
           row_range = ws[1:1]
           cols=0
           fields_py=[]
           for cells in row_range[:ws.max_column]:
               cols=cols+1
               comment = cells.comment
               if comment:print comment
               if len(cells.value)>0: #is types.UnicodeType:
                   field=cells.value.replace("\n","")
                   field_py=lazy_pinyin(field,errors='replace')
                   py=""
                   for p in field_py:
                       py=py + p #.capitalize()
                   print "   ",field, py
                   fields_py.append(py)
                   attributes=""
                   if cols<len(ts) :
                       if len(ts[cols])>0:attributes=":"+ts[cols]
                   if py=="riqi" or '日' in field: attributes=":date"
                   if field in sheetlist:attributes=":belongs_to"



                   fw.write(" "+py+attributes)
                   fy.write("\n  "+py+": '"+field+"'")
           fw.write("  --skip-stylesheets\n")

           for row in sheet.rows:

               i=0
               data_s=sheetname_py.capitalize()+".create("
               for cell in row:
                   data_s=data_s+fields_py[i]+": '"+ str(cell.value)+"', "
                   i+=1
               data_s=data_s[:-2]+")\n"
               fw_create.write(data_s)

fw.write("rails db:migrate\n")
fw.write("cp  ../Create_data.rb db/seeds.rb \n")
fw.write("rails db:seed\n")
fw.write("sed -i '1a\  get \"home/index\"' config/routes.rb\n")
fw.write("sed -i '"+str(len(views)+3)+"a\  root \"home#index\"' config/routes.rb\n")
#fw.write("cd ..\n")
fw.write("cp  ../index.html.erb ")
#fw.write(app_py)
fw.write("app/views/home/index.html.erb\n")

fw.write("sed -i \"7a\gem \'twitter-bootstrap-rails\'\" Gemfile\n")
fw.write("sed -i \"8a\gem \'therubyracer\'\" Gemfile\n")
fw.write("sed -i \"9a\gem \'less-rails\'\" Gemfile\n")
fw.write("sed -i \"7a\gem \'devise\'\" Gemfile\n")
fw.write("bundle install\n")

fw.write("rails g bootstrap:install\n")

for l in views:
    fw.write("rails g bootstrap:themed "+l+" -f\n")

fw.write("cp -f ../en.yml ")
#fw.write(app_py)
fw.write("config/locales/en.yml\n")
dirs= "app/views/"

for l in views:
    f=dirs+l+"/_form.html.erb"
    fw.write("sed -i 's/f.label/f.label t/' "+f+"\n")

#fw.write("mv app/assets/stylesheets/application.css app/assets/stylesheets/application.scss\n")
#fw.write("sed -i '15a\// \"bootstrap-sprockets\" must be imported before \"bootstrap\" and \"bootstrap/variables\"' app/assets/stylesheets/application.scss\n")
#fw.write("sed -i '16a\@import \"bootstrap-sprockets\";' app/assets/stylesheets/application.scss\n")
#fw.write("sed -i '17a\@import \"bootstrap\";' app/assets/stylesheets/application.scss\n")




#fw.write("bundle install\n")

fw.write("rails g devise:install\n")
fw.write("rails g devise user\n")
fw.write("rails db:migrate\n")
fw.write("rails server\n")
fw.write("firefox localhost:3000\n")
fw.close()
fy.close()
f_index.close()
